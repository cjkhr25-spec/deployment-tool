import os
import json
import shutil
import zipfile
import webbrowser
import threading
import csv
from datetime import datetime
from ftplib import FTP, error_perm
from tkinter import Tk, Label, Button, Scrollbar, filedialog, messagebox, Text, Frame
from tkinter import VERTICAL, RIGHT, Y, BOTH, END, LEFT, NORMAL, DISABLED

# 외부 라이브러리 연동
import openpyxl 
from tkinter.ttk import Treeview, Style, Progressbar

# 고정 환경 설정 변수
CONFIG_FILE = "config.json"
FTP_HOST = "183.110.210.22"
REMOTE_ZIP_DIR = "/New_WWW/webdata1/Cache/flash/Multiplayer/"
REMOTE_FOLDER_DIR = "/New_WWW/webdata1/Cache/webflash/Multiplayer/"
REVIEW_FILE = "review_urls.txt"
RESULT_CSV_FILE = "deployment_result.csv"

class DeploymentToolApp:
    def __init__(self, root):
        self.root = root
        self.root.title("체험웹 배포 도구")
        self.root.geometry("550x800")
        self.root.resizable(False, False)

        # 시스템 내부 데이터 관리 변수
        self.selected_files = [] 
        self.excel_path = ""      
        self.ftp_user = "" 
        self.ftp_pass = "" 
        
        # 파일 로그를 추적하기 위한 핸들 경로
        self.log_file_path = None

        # UI 기본 스타일 구성
        self.style = Style()
        self.style.theme_use("clam")

        self.create_widgets()
        
        # 초기 설정 및 로그 인프라 빌드
        self.initialize_config_and_logs()

    def create_widgets(self):
        # ------------------------------------------------------------
        # 프로그레스바 및 상단 현재 처리 상태 안내 영역
        # ------------------------------------------------------------
        self.progress_frame = Frame(self.root)
        self.progress_frame.pack(fill=BOTH, padx=20, pady=(15, 5))
        
        self.label_progress = Label(self.progress_frame, text="대기 중 (파일을 추가해 주세요)", font=("맑은 고딕", 9), fg="#555555")
        self.label_progress.pack(anchor="w")
        
        self.progress_bar = Progressbar(self.progress_frame, orient="horizontal", mode="determinate")
        self.progress_bar.pack(fill=BOTH, expand=True, pady=(3, 0))
        self.progress_bar["value"] = 0 

        # ------------------------------------------------------------
        # 1. ZIP 파일 추가 섹션
        # ------------------------------------------------------------
        self.label_title = Label(self.root, text="[ ZIP 파일 추가 ]", font=("맑은 고딕", 11, "bold"))
        self.label_title.pack(anchor="w", padx=20, pady=(10, 5))

        self.zip_btn_frame = Frame(self.root)
        self.zip_btn_frame.pack(anchor="w", padx=20, pady=(0, 5))

        self.btn_add_zip = Button(self.zip_btn_frame, text="ZIP 파일 선택", command=self.add_zip_files, bg="#f0f0f0", relief="groove", width=12)
        self.btn_add_zip.pack(side=LEFT) 

        self.btn_delete_selected = Button(self.zip_btn_frame, text="선택 삭제", command=self.delete_selected_files, bg="#ffcccc", relief="groove", width=10)
        self.btn_delete_selected.pack(side=LEFT, padx=5)

        self.btn_delete_all = Button(self.zip_btn_frame, text="전체 삭제", command=self.delete_all_files, bg="#ff9999", relief="groove", width=10)
        self.btn_delete_all.pack(side=LEFT, padx=5)

        self.btn_open_web = Button(self.zip_btn_frame, text="선택 검수 열기", command=self.open_review_page, bg="#f0f0f0", relief="groove", width=12)
        self.btn_open_web.pack(side=LEFT, padx=5)

        # ------------------------------------------------------------
        # 2. 파일 목록 표시 영역 (Treeview)
        # ------------------------------------------------------------
        self.tree_frame = Label(self.root)
        self.tree_frame.pack(fill=BOTH, padx=20, pady=(0, 15))

        self.tree_scroll = Scrollbar(self.tree_frame, orient=VERTICAL)
        self.tree_scroll.pack(side=RIGHT, fill=Y)

        self.file_tree = Treeview(self.tree_frame, columns=("filename", "mcode", "status"), show="headings", yscrollcommand=self.tree_scroll.set, height=6)
        self.file_tree.pack(fill=BOTH, expand=True, side=LEFT)
        self.tree_scroll.config(command=self.file_tree.yview)

        self.file_tree.heading("filename", text="파일명", anchor="w")
        self.file_tree.heading("mcode", text="mcode", anchor="center")
        self.file_tree.heading("status", text="상태", anchor="center")
        
        self.file_tree.column("filename", width=220, anchor="w")
        self.file_tree.column("mcode", width=130, anchor="center")
        self.file_tree.column("status", width=120, anchor="center")

        self.file_tree.bind("<Double-1>", self.open_review_page)

        # ------------------------------------------------------------
        # 3. 배포 경로 엑셀 선택 섹션
        # ------------------------------------------------------------
        self.label_excel = Label(self.root, text="[ 경로정보.xlsx 선택 ]", font=("맑은 고딕", 11, "bold"))
        self.label_excel.pack(anchor="w", padx=20, pady=(5, 5))

        self.btn_add_excel = Button(self.root, text="엑셀 파일 선택하기", command=self.add_excel_file, bg="#f0f0f0", relief="groove", width=20)
        self.btn_add_excel.pack(anchor="w", padx=20, pady=(0, 5))

        self.label_excel_path = Label(self.root, text="선택된 엑셀 파일이 없습니다.", font=("맑은 고딕", 9), fg="#666666", wraplength=500, justify="left")
        self.label_excel_path.pack(anchor="w", padx=20, pady=(0, 10))

        # ------------------------------------------------------------
        # 성공 대상 일괄 사후 검수 확장 기능 컴포넌트 섹션
        # ------------------------------------------------------------
        self.review_frame = Frame(self.root)
        self.review_frame.pack(anchor="w", padx=20, pady=(0, 15))
        
        self.label_review_title = Label(self.review_frame, text="[ 배포 성공 콘텐츠 일괄 검수 기능 ]", font=("맑은 고딕", 9, "bold"), fg="#2c3e50")
        self.label_review_title.pack(anchor="w", pady=(0, 3))
        
        self.btn_copy_urls = Button(self.review_frame, text="📋 성공 URL 전체 복사", command=self.copy_success_urls_to_clipboard, bg="#e8f4f8", relief="groove", font=("맑은 고딕", 9), width=22)
        self.btn_copy_urls.pack(side=LEFT)
        
        self.btn_open_all_success = Button(self.review_frame, text="🚀 성공 브라우저 일괄 오픈", command=self.open_all_success_pages, bg="#eafaf1", relief="groove", font=("맑은 고딕", 9), width=22)
        self.btn_open_all_success.pack(side=LEFT, padx=10)

        # ------------------------------------------------------------
        # 5. 실행 및 하단 로그 출력 영역
        # ------------------------------------------------------------
        self.btn_start = Button(self.root, text="배포 시작", command=self.start_deployment_thread, bg="#e1e1e1", font=("맑은 고딕", 10, "bold"), width=15)
        self.btn_start.pack(pady=(0, 10))

        self.label_log = Label(self.root, text="진행상황 및 로그", font=("맑은 고딕", 10, "bold"))
        self.label_log.pack(anchor="w", padx=20)

        self.log_text = Text(self.root, height=8, font=("Consolas", 9), bg="#fafafa", highlightthickness=1, highlightbackground="#cccccc")
        self.log_text.pack(fill=BOTH, padx=20, pady=(0, 20))
        self.log_text.config(state=DISABLED)

    def initialize_config_and_logs(self):
        if not os.path.exists("logs"):
            os.makedirs("logs")

        timestamp = datetime.now().strftime("%Y-%m-%d")
        self.log_file_path = f"logs/{timestamp}.log"

        if not os.path.exists(CONFIG_FILE):
            default_config = {
                "excel_path": "",
                "ftp_user": "your_ftp_username",
                "ftp_pass": "your_ftp_password"
            }
            try:
                with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                    json.dump(default_config, f, ensure_ascii=False, indent=4)
                self.write_log("INFO", "config.json 파일이 존재하지 않아 기본 파일로 자동 생성했습니다.")
            except Exception as e:
                print(f"초기 설정 파일 생성 오류: {e}")
                
        self.load_config()

    def write_log(self, level, message):
        full_msg = f"[{level}] {message}"
        self.root.after(0, lambda: self._safe_write_text_ui(full_msg))

        if self.log_file_path:
            try:
                with open(self.log_file_path, "a", encoding="utf-8") as f:
                    f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {full_msg}\n")
            except Exception:
                pass

    def _safe_write_text_ui(self, text):
        self.log_text.config(state=NORMAL)
        self.log_text.insert(END, text + "\n")
        self.log_text.see(END)
        self.log_text.config(state=DISABLED)

    # ------------------------------------------------------------
    # FTP 비즈니스 네트워크 통신 코어 (재접속 및 딥 메이크 폴더 탑재)
    # ------------------------------------------------------------

    def ftp_connect(self):
        if not self.ftp_user or not self.ftp_pass:
            self.write_log("ERROR", "config.json 파일에 FTP 계정 정보가 누락되었습니다.")
            return None
        try:
            ftp = FTP()
            ftp.connect(FTP_HOST, 21, timeout=20)
            ftp.login(self.ftp_user, self.ftp_pass)
            ftp.encoding = "utf-8"
            return ftp
        except Exception as e:
            self.write_log("ERROR", f"FTP 최초 연결 에러 ({FTP_HOST}): {e}")
            return None

    def execute_ftp_with_retry(self, ftp_session, action_func, *args, **kwargs):
        try:
            return action_func(ftp_session, *args, **kwargs), ftp_session
        except Exception as first_error:
            self.write_log("WARNING", f"원격 명령 수행 에러 발생 ({first_error}). 세션을 복구하여 1회 재시도합니다.")
            
            try:
                ftp_session.close()
            except Exception:
                pass
                
            new_session = self.ftp_connect()
            if not new_session:
                raise Exception(f"네트워크 단절로 인한 FTP 재접속 최종 실패 (상세사유: {first_error})")
                
            try:
                result = action_func(new_session, *args, **kwargs)
                return result, new_session
            except Exception as second_error:
                raise Exception(f"재접속 후 최종 가동 실패 (최종사유: {second_error})")

    def _make_directory_p(self, ftp, remote_dir_path):
        parts = [p for p in remote_dir_path.replace("\\", "/").split("/") if p]
        current_dir = ""
        if remote_dir_path.startswith("/"):
            current_dir = "/"

        for part in parts:
            current_dir = os.path.join(current_dir, part).replace("\\", "/")
            try:
                ftp.cwd(current_dir)
            except error_perm:
                try:
                    ftp.mkd(current_dir)
                    ftp.cwd(current_dir)
                except Exception as e:
                    raise Exception(f"원격 mkdir -p 폴더 생성 오류 ({current_dir}): {e}")

    # ------------------------------------------------------------
    # FTP 원격 입출력 서브 태스크 모듈 (자원 핸들 누수 완전 폐쇄)
    # ------------------------------------------------------------

    def extract_zip(self, zip_path, extract_to):
        try:
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(extract_to)
            return True
        except zipfile.BadZipFile:
            raise Exception("손상된 ZIP 파일입니다.")
        except Exception as e:
            raise Exception(f"압축 해제 실패 : {e}")

    def _core_upload_zip(self, ftp, local_zip_path, filename):
        ftp.cwd(REMOTE_ZIP_DIR)
        file_handle = open(local_zip_path, 'rb')
        try:
            ftp.storbinary(f'STOR {filename}', file_handle)
        finally:
            file_handle.close()

    def _core_upload_directory_recursive(self, ftp, local_dir_path, remote_dir_path):
        self._make_directory_p(ftp, remote_dir_path)

        for name in os.listdir(local_dir_path):
            local_item_path = os.path.join(local_dir_path, name)
            remote_item_path = remote_dir_path.rstrip('/') + '/' + name

            if os.path.isdir(local_item_path):
                self._core_upload_directory_recursive(ftp, local_item_path, remote_item_path)
            else:
                ftp.cwd(remote_dir_path)
                file_handle = open(local_item_path, 'rb')
                try:
                    ftp.storbinary(f'STOR {name}', file_handle)
                finally:
                    file_handle.close()

    # ------------------------------------------------------------
    # 배포 백그라운드 스레드 제어부 (UI 프리징 완전 해결)
    # ------------------------------------------------------------

    def start_deployment_thread(self):
        if not self.selected_files:
            messagebox.showwarning("경고", "배포할 ZIP 파일을 먼저 추가해 주세요.")
            return
        if not self.excel_path:
            messagebox.showwarning("경고", "경로정보 엑셀 파일을 먼저 선택해 주세요.")
            return

        self.btn_start.config(state=DISABLED)
        
        deploy_thread = threading.Thread(target=self.start_deployment, daemon=True)
        deploy_thread.start()

    def start_deployment(self):
        self.write_log("INFO", "▶ 통합 자동 배포 프로세스를 개시합니다.\n" + "═"*40)
        
        if not self.find_mcode_from_excel():
            self.write_log("ERROR", "❌ 엑셀 매칭 단계에 치명적 오류가 확인되어 작업을 중단합니다.")
            self.root.after(0, lambda: self.btn_start.config(state=NORMAL))
            return

        ftp = self.ftp_connect()
        if not ftp:
            self.write_log("ERROR", "❌ 초기 FTP 서버 접속에 실패하여 원격 전송 파이프라인을 취소합니다.")
            self.root.after(0, lambda: self.btn_start.config(state=NORMAL))
            return

        success_list = []
        failed_list = []
        
        review_urls_data = []
        csv_report_data = []

        for idx, file_obj in enumerate(self.selected_files):
            filename = file_obj["filename"]
            local_path = file_obj["path"]
            temp_extract_dir = local_path + "_temp_dir"
            current_time_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            if file_obj["status"] != "매칭완료":
                self.write_log("WARNING", f"⏩ [스킵] {filename} - 매칭 결과가 누락되어 배포 라인업에서 제외됩니다.")
                continue

            progress_text = f"[{idx + 1} / {len(self.selected_files)}] {filename} 업로드 중..."
            self.root.after(0, lambda p=progress_text: self.label_progress.config(text=p, fg="#0066cc"))

            try:
                # 1. 로컬 임시 폴더에 압축 해제
                self.update_status(filename, "압축해제")
                self.extract_zip(local_path, temp_extract_dir)

                # 2. 압축 해제 후 생성된 알맹이(최상위 단일 폴더) 탐색 검증 로직
                target_folder_name = os.path.splitext(filename)[0]
                full_target_folder_path = temp_extract_dir

                if not os.path.isdir(full_target_folder_path):
                    raise Exception("압축 해제 후 최상위 항목이 폴더 형식이 아닌 파일 구조입니다.")

                # 3. 원격 ZIP 아카이브 방 전송
                self.update_status(filename, "업로드중")
                _, ftp = self.execute_ftp_with_retry(ftp, self._core_upload_zip, local_path, filename)

                # 4. 탐색된 실제 콘텐츠 폴더 기반으로 FTP 원격지 타겟 경로 동적 가공 및 재귀 전송
                target_remote_path = REMOTE_FOLDER_DIR.rstrip('/') + '/' + target_folder_name
                _, ftp = self.execute_ftp_with_retry(ftp, self._core_upload_directory_recursive, full_target_folder_path, target_remote_path)

                # 성공 처리 기록 보관
                self.update_status(filename, "완료")
                success_list.append(filename)
                
                generated_url = f"https://test-web.milkt.co.kr/HtmlPlayer?id={file_obj['mcode']}"
                review_urls_data.append(generated_url)
                
                csv_report_data.append([filename, file_obj['mcode'], "성공", "", current_time_str])
                self.write_log("INFO", f"✔ [성공] {filename} -> 원격 /{target_folder_name} 배포 완료")

            except Exception as file_error:
                # 개별 파일 실패 시 예외를 캐치하여 프로그램이 죽지 않고 다음 파일 작업을 연속 처리하도록 구현
                self.update_status(filename, "실패")
                error_message = str(file_error)
                failed_list.append((filename, error_message))
                
                csv_report_data.append([filename, file_obj['mcode'], "실패", error_message, current_time_str])
                self.write_log("ERROR", f"❌ [실패] {filename} - 원인: {error_message}")

            finally:
                # 생성된 로컬 임시 작업 공간 안전하게 휘발 삭제 마감
                if os.path.exists(temp_extract_dir):
                    try:
                        shutil.rmtree(temp_extract_dir)
                    except Exception as e:
                        print(f"임시 폴더 삭제 에러: {e}")

            current_percentage = int(((idx + 1) / len(self.selected_files)) * 100)
            self.root.after(0, lambda v=current_percentage: self.progress_bar.config(value=v))

        if ftp:
            try:
                ftp.quit()
            except Exception:
                try:
                    ftp.close()
                except Exception:
                    pass

        # 종합 리포트 출력 마감
        self.root.after(0, lambda: self.label_progress.config(text="모든 배포 작업이 마무리되었습니다.", fg="#27ae60"))
        
        if review_urls_data:
            try:
                with open(REVIEW_FILE, "w", encoding="utf-8") as rf:
                    rf.write("\n".join(review_urls_data))
                self.write_log("INFO", f"💾 시스템: 성공 콘텐츠 검수 URL 목록이 '{REVIEW_FILE}'로 자동 생성되었습니다.")
                
                self.root.clipboard_clear()
                self.root.clipboard_append("\n".join(review_urls_data))
                self.root.update()
                self.write_log("INFO", "📋 시스템: 성공 건에 대한 모든 검수 URL이 클립보드에 자동 복사되었습니다.")
            except Exception as fe:
                self.write_log("ERROR", f"⚠️ 검수 URL 텍스트 파일 저장 중 예외 발생: {fe}")

        try:
            with open(RESULT_CSV_FILE, "w", encoding="utf-8-sig", newline="") as cf:
                writer = csv.writer(cf)
                writer.writerow(["파일명", "mcode", "배포결과", "실패사유", "배포시간"])
                writer.writerows(csv_report_data)
            self.write_log("INFO", f"💾 시스템: 종합 배포 결과 명세서가 '{RESULT_CSV_FILE}'로 내보내졌습니다.")
        except Exception as ce:
            self.write_log("ERROR", f"⚠️ 종합 결과 분석 CSV 파일 생성 실패: {ce}")
        
        summary_report = [
            "\n================================",
            "배포 완료",
            f"\n성공 개수 : {len(success_list)}건",
            f"실패 개수 : {len(failed_list)}건",
            "\n실패 사유"
        ]
        if failed_list:
            for item, reason in failed_list:
                summary_report.append(f"- {item} ➔ {reason}")
        else:
            summary_report.append("- 없음")
        summary_report.append("================================")
        
        self.write_log("INFO", "\n".join(summary_report))
        self.root.after(0, lambda: self.btn_start.config(state=NORMAL))

    # ------------------------------------------------------------
    # 사후 검수 부가 기능 이벤트 핸들러
    # ------------------------------------------------------------

    def copy_success_urls_to_clipboard(self):
        success_urls = []
        for file_obj in self.selected_files:
            if file_obj["status"] == "완료" and file_obj["mcode"]:
                url = f"https://test-web.milkt.co.kr/HtmlPlayer?id={file_obj['mcode']}"
                success_urls.append(url)

        if not success_urls:
            messagebox.showwarning("안내", "배포 완료(성공) 상태의 유효한 mcode 콘텐츠가 없습니다.")
            return

        self.root.clipboard_clear()
        self.root.clipboard_append("\n".join(success_urls))
        self.root.update()
        messagebox.showinfo("복사 완료", "성공한 콘텐츠의 검수 URL이 클립보드에 일괄 복사되었습니다.")

    def open_all_success_pages(self):
        success_count = 0
        for file_obj in self.selected_files:
            if file_obj["status"] == "완료" and file_obj["mcode"]:
                url = f"https://test-web.milkt.co.kr/HtmlPlayer?id={file_obj['mcode']}"
                try:
                    webbrowser.open_new_tab(url)
                    success_count += 1
                except Exception as e:
                    self.write_log("WARNING", f"브라우저 호출 중 예외 발생: {e}")

        if success_count == 0:
            messagebox.showwarning("안내", "일괄 오픈할 배포 성공 완료 항목이 목록에 존재하지 않습니다.")

    # ------------------------------------------------------------
    # UI 이벤트 및 유틸리티 함수
    # ------------------------------------------------------------

    def open_review_page(self, event=None):
        selected_items = self.file_tree.selection()
        if not selected_items:
            messagebox.showinfo("안내", "검수할 콘텐츠를 목록에서 선택해주세요.")
            return

        for item in selected_items:
            values = self.file_tree.item(item, "values")
            filename, mcode = values[0], values[1]

            if not mcode or mcode.strip() == "":
                self.write_log("WARNING", f"❌ [오류] 먼저 mcode를 조회해주세요. (파일명: {filename})")
                continue

            try:
                review_url = f"https://test-web.milkt.co.kr/HtmlPlayer?id={mcode}"
                webbrowser.open_new_tab(review_url)
            except Exception as e:
                self.write_log("ERROR", f"브라우저 연동 실패: {e}")

    def add_zip_files(self):
        try:
            files = filedialog.askopenfilenames(title="배포할 ZIP 파일들을 선택하세요", filetypes=[("ZIP 파일", "*.zip"), ("모든 파일", "*.*")])
            if files:
                for file_path in files:
                    existing_paths = [file_obj["path"] for file_obj in self.selected_files]
                    if file_path not in existing_paths:
                        file_name = os.path.basename(file_path)
                        file_object = {"path": file_path, "filename": file_name, "mcode": "", "status": "대기"}
                        self.selected_files.append(file_object)
                        self.file_tree.insert("", END, values=(file_name, "", "대기"))
                self.write_log("INFO", f"ZIP 파일 {len(files)}개가 목록에 추가되었습니다.")
        except Exception as e:
            messagebox.showerror("오류", f"파일 시스템 스캔 실패:\n{e}")

    def update_status(self, filename, new_status):
        for file_obj in self.selected_files:
            if file_obj["filename"] == filename:
                file_obj["status"] = new_status
                break
        self.root.after(0, lambda: self._safe_update_tree_status(filename, new_status))

    def _safe_update_tree_status(self, filename, new_status):
        for item in self.file_tree.get_children():
            values = self.file_tree.item(item, "values")
            if values[0] == filename:
                self.file_tree.item(item, values=(values[0], values[1], new_status))
                break

    def update_mcode(self, filename, mcode):
        for file_obj in self.selected_files:
            if file_obj["filename"] == filename:
                file_obj["mcode"] = mcode
                break
        self.root.after(0, lambda: self._safe_update_tree_mcode(filename, mcode))

    def _safe_update_tree_mcode(self, filename, mcode):
        for item in self.file_tree.get_children():
            values = self.file_tree.item(item, "values")
            if values[0] == filename:
                self.file_tree.item(item, values=(values[0], mcode, values[2]))
                break

    def get_filename_from_path(self, path_string):
        if not path_string: return ""
        return path_string.replace("\\", "/").split("/")[-1]

    def find_mcode_from_excel(self):
        wb = None
        try:
            wb = openpyxl.load_workbook(self.excel_path, data_only=True)
            sheet = wb.active 
            mcode_col_idx, path_col_idx = None, None
            
            for col_idx in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col_idx).value
                if cell_value == "mcode": mcode_col_idx = col_idx
                elif cell_value == "경로": path_col_idx = col_idx

            if mcode_col_idx is None or path_col_idx is None:
                self.write_log("WARNING", "⚠️ [형식오류] 엑셀에 'mcode' 또는 '경로' 컬럼명이 존재하지 않습니다.")
                return False

            for file_obj in self.selected_files:
                target_zip_name = file_obj["filename"] 
                is_matched = False 
                
                for row_idx in range(2, sheet.max_row + 1):
                    excel_path_value = sheet.cell(row=row_idx, column=path_col_idx).value
                    excel_mcode_value = sheet.cell(row=row_idx, column=mcode_col_idx).value
                    
                    if excel_path_value:
                        excel_filename = self.get_filename_from_path(str(excel_path_value))
                        if excel_filename.lower() == target_zip_name.lower():
                            mcode_result = str(excel_mcode_value) if excel_mcode_value else ""
                            self.update_mcode(target_zip_name, mcode_result)
                            self.update_status(target_zip_name, "매칭완료")
                            is_matched = True
                            break 
                
                if not is_matched:
                    self.update_status(target_zip_name, "경로없음")
                    self.write_log("WARNING", f"[경로없음] 엑셀 매칭 정보 없음 - ZIP: {target_zip_name}")
            return True
        except Exception as e:
            self.write_log("ERROR", f"⚠️ [엑셀오류] 파일 로드 실패: {e}")
            return False
        finally:
            if wb:
                try:
                    wb.close()
                except Exception:
                    pass

    def delete_selected_files(self):
        selected_items = self.file_tree.selection()
        if not selected_items: return
        for item in selected_items:
            values = self.file_tree.item(item, "values")
            file_name = values[0] 
            for file_obj in self.selected_files:
                if file_obj["filename"] == file_name:
                    self.selected_files.remove(file_obj)
                    break
            self.file_tree.delete(item)

    def delete_all_files(self):
        if not self.selected_files: return
        if messagebox.askyesno("확인", "목록에 있는 모든 파일을 삭제하시겠습니까?"):
            self.selected_files.clear() 
            for item in self.file_tree.get_children(): self.file_tree.delete(item)
            self.write_log("INFO", "모든 파일 목록이 초기화되었습니다.")

    def add_excel_file(self):
        try:
            file = filedialog.askopenfilename(title="경로정보 엑셀 파일을 선택하세요", filetypes=[("엑셀 파일", "*.xlsx *.xls"), ("모든 파일", "*.*")])
            if file:
                self.excel_path = file
                self.label_excel_path.config(text=file, fg="#0066cc")
                self.write_log("INFO", f"연동용 경로 엑셀 지정 완료: {os.path.basename(file)}")
                self.save_config()
        except Exception as e:
            messagebox.showerror("오류", f"엑셀 선택 차단 예외: {e}")

    def save_config(self):
        try:
            config_data = {"excel_path": self.excel_path, "ftp_user": self.ftp_user, "ftp_pass": self.ftp_pass}
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(config_data, f, ensure_ascii=False, indent=4)
        except Exception as e:
            print(f"설정 입출력 에러: {e}")

    def load_config(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    config_data = json.load(f)
                    self.excel_path = config_data.get("excel_path", "")
                    self.ftp_user = config_data.get("ftp_user", "")
                    self.ftp_pass = config_data.get("ftp_pass", "")
                    if self.excel_path and os.path.exists(self.excel_path):
                        self.label_excel_path.config(text=self.excel_path, fg="#0066cc")
            except Exception as e:
                print(f"설정 복원 에러: {e}")

if __name__ == "__main__":
    root = Tk()
    app = DeploymentToolApp(root)
    root.mainloop()