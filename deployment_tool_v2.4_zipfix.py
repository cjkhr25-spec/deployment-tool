import os
import json
import time
import shutil
import zipfile
import webbrowser
import threading
import traceback
import csv
from datetime import datetime
from ftplib import FTP, error_perm
from tkinter import Tk, Label, Button, Scrollbar, filedialog, messagebox, Text, Frame
from tkinter import VERTICAL, RIGHT, Y, BOTH, END, LEFT, NORMAL, DISABLED

# 외부 라이브러리 연동
import openpyxl 
from tkinter.ttk import Treeview, Style, Progressbar

# 고정 환경 설정 변수
CONFIG_FILE = "config.json"
FTP_HOST = "183.110.210.7"
FTP_PORT = 7021

# 기준이 되는 사내 원격 절대 루트 상위 디렉토리
REMOTE_ROOT_DIR = "/New_WWW/webdata1/Cache"

REVIEW_FILE = "review_urls.txt"
RESULT_CSV_FILE = "deployment_result.csv"

class DeploymentToolApp:
    def __init__(self, root):
        self.root = root
        self.root.title("🚀 체험웹 자동 배포 시스템 v3.0")
        self.root.geometry("920x860")
        self.root.minsize(900, 820)
        self.root.configure(bg="#F3F4F6")

        # ------------------------------------------
        # 프로그램 정보
        # ------------------------------------------
        self.version = "v3.0"

        # ------------------------------------------
        # 내부 데이터
        # ------------------------------------------
        self.selected_files = []
        self.excel_path = ""

        self.ftp_user = ""
        self.ftp_pass = ""
        self.remote_root = ""

        self.log_file_path = None

        self.total_files = 0
        self.success_count = 0
        self.fail_count = 0

        self.current_filename = ""
        self.current_path = ""

        self.start_time = None

        # ------------------------------------------
        # 스타일
        # ------------------------------------------
        self.style = Style()
        self.style.theme_use("clam")

        self.style.configure(
            ".",
            background="#F3F4F6",
            foreground="#1F2937",
            font=("맑은 고딕", 9)
        )

        self.style.configure(
            "Treeview",
            background="#FFFFFF",
            fieldbackground="#FFFFFF",
            foreground="#1D1D1F",
            rowheight=38,
            font=("맑은 고딕", 10),
            borderwidth=0,
            relief="flat"
        )

        self.style.configure(
            "Treeview.Heading",
            background="#F9FAFB",
            foreground="#6B7280",
            font=("맑은 고딕", 10, "bold"),
            borderwidth=0,
            relief="flat",
            padding=(0, 12)
        )

        self.style.map(
            "Treeview",
            background=[("selected", "#EAF2FF")],
            foreground=[("selected", "#1D1D1F")]
        )

        self.style.configure(
            "Blue.Horizontal.TProgressbar",
            troughcolor="#E5E7EB",
            background="#2563EB",
            thickness=14,
            borderwidth=0
        )

        self.button_font = ("맑은 고딕", 9)

        try:
            self.log_font = ("D2Coding", 10)
        except Exception:
            self.log_font = ("Consolas", 10)

        self.create_widgets()

        self.file_tree.tag_configure("ready", foreground="#6B7280")
        self.file_tree.tag_configure("working", foreground="#2563EB")
        self.file_tree.tag_configure("success", foreground="#16A34A")
        self.file_tree.tag_configure("fail", foreground="#DC2626")

        self.initialize_config_and_logs()

    def create_widgets(self):

        # =====================================================
        # 공통 버튼 옵션
        # =====================================================

        btn_opts = {
            "font": ("맑은 고딕", 10),
            "relief": "flat",
            "cursor": "hand2",
            "padx": 18,
            "pady": 8,
            "bd": 1,
            "highlightthickness": 1,
            "highlightbackground": "#D1D5DB",
            "highlightcolor": "#D1D5DB"
        }

        # =====================================================
        # 메인 컨테이너
        # =====================================================

        self.main_frame = Frame(
            self.root,
            bg="#F3F4F6",
        )
        self.main_frame.pack(fill=BOTH, expand=True, padx=24, pady=20)

        # =====================================================
        # 프로그램 헤더
        # =====================================================

        self.header_frame = Frame(
            self.main_frame,
            bg="#FFFFFF",
            bd=0,
            relief="flat"
        )
        self.header_frame.pack(fill="x", pady=(0, 16))

        self.label_program = Label(
            self.header_frame,
            text="HTML 자동 배포 시스템",
            font=("맑은 고딕", 20, "bold"),
            fg="#1F2937",
            bg="#FFFFFF"
        )
        self.label_program.pack(anchor="w", padx=20, pady=(16, 2))

        self.label_version = Label(
            self.header_frame,
            text="MILKT Internal Deployment Tool",
            font=("맑은 고딕", 10),
            fg="#8E8E93",
            bg="#FFFFFF"
        )
        self.label_version.pack(anchor="w", padx=20, pady=(0, 16))

        # =====================================================
        # 진행상황 카드
        # =====================================================

        self.progress_card = Frame(
            self.main_frame,
            bg="#FFFFFF",
            bd=0,
            relief="flat"
        )
        self.progress_card.pack(fill="x", pady=(0, 18))

        self.label_progress = Label(
            self.progress_card,
            text="대기 중",
            font=("맑은 고딕", 13, "bold"),
            fg="#111827",
            bg="#FFFFFF"
        )
        self.label_progress.pack(anchor="w", padx=20, pady=(16, 2))

        self.label_current_file = Label(
            self.progress_card,
            text="현재 작업 : 없음",
            font=("맑은 고딕", 10),
            fg="#4B5563",
            bg="#FFFFFF"
        )
        self.label_current_file.pack(anchor="w", padx=20)

        self.label_current_path = Label(
            self.progress_card,
            text="",
            font=("Consolas", 9),
            fg="#9AA0A6",
            bg="#FFFFFF"
        )
        self.label_current_path.pack(anchor="w", padx=20, pady=(0, 10))
    
        self.label_count = Label(
            self.progress_card,
            text="진행 : 0 / 0",
            font=("맑은 고딕", 10),
            fg="#6B7280",
            bg="#FFFFFF"
        )
        self.label_count.pack(anchor="e", padx=20, pady=(0, 10))

        self.progress_bar = Progressbar(
            self.progress_card,
            orient="horizontal",
            mode="determinate",
            style="Blue.Horizontal.TProgressbar"
        )
        self.progress_bar.pack(fill="x", padx=20)

        self.progress_bar["value"] = 0

        self.label_percent = Label(
            self.progress_card,
            text="0 %",
            font=("맑은 고딕", 12, "bold"),
            fg="#111827",
            bg="#FFFFFF"
        )
        self.label_percent.pack(anchor="e", padx=20, pady=(6, 16))

        # =====================================================
        # ZIP 파일 카드
        # =====================================================

        self.zip_card = Frame(
            self.main_frame,
            bg="#FFFFFF",
            bd=0,
            relief="flat"
        )
        self.zip_card.pack(fill="x", pady=(0,18))

        # ---------------- 제목 ----------------

        self.label_title = Label(
            self.zip_card,
            text="ZIP 파일",
            font=("맑은 고딕", 14, "bold"),
            bg="#FFFFFF",
            fg="#1F2937"
        )
        self.label_title.pack(anchor="w", padx=20, pady=(18, 10))

        # ---------------- 버튼 영역 ----------------

        self.zip_btn_frame = Frame(
            self.zip_card,
            bg="#FFFFFF"
        )
        self.zip_btn_frame.pack(fill="x", padx=24, pady=(0, 6))

        self.btn_add_zip = Button(
            self.zip_btn_frame,
            text="＋ ZIP 추가",
            command=self.add_zip_files,
            bg="#F3F4F6",
            fg="#1F2937",
            activebackground="#E5E7EB",
            activeforeground="white",
            width=13,
            **btn_opts
        )
        self.btn_add_zip.pack(side=LEFT, padx=(0, 8))

        self.btn_delete_selected = Button(
            self.zip_btn_frame,
            text="선택 삭제",
            command=self.delete_selected_files,
            bg="#F3F4F6",
            fg="#374151",
            activebackground="#E5E7EB",
            width=11,
            **btn_opts
        )
        self.btn_delete_selected.pack(side=LEFT, padx=(0, 8))

        self.btn_delete_all = Button(
            self.zip_btn_frame,
            text="전체 삭제",
            command=self.delete_all_files,
            bg="#F3F4F6",
            fg="#DC2626",
            activebackground="#DC2626",
            activeforeground="white",
            width=11,
            **btn_opts
        )
        self.btn_delete_all.pack(side=LEFT, padx=(0, 8))

        self.btn_open_web = Button(
            self.zip_btn_frame,
            text="검수 페이지",
            command=self.open_review_page,
            bg="#F3F4F6",
            fg="#1F2937",
            activebackground="#E5E7EB",
            activeforeground="white",
            width=12,
            **btn_opts
        )
        self.btn_open_web.pack(side="right")

        # ---------------- 목록 ----------------

        self.tree_frame = Frame(
            self.zip_card,
            bg="#FFFFFF"
        )
        self.tree_frame.pack(fill=BOTH, expand=True, padx=24, pady=(16, 22))

        self.tree_scroll = Scrollbar(
            self.tree_frame,
            orient=VERTICAL
        )
        self.tree_scroll.pack(side=RIGHT, fill=Y)

        self.file_tree = Treeview(
            self.tree_frame,
            columns=("filename", "mcode", "status"),
            show="headings",
            height=8,
            yscrollcommand=self.tree_scroll.set
        )

        self.file_tree.pack(
            side=LEFT,
            fill=BOTH,
            expand=True
        )

        self.tree_scroll.config(command=self.file_tree.yview)

        self.file_tree.heading(
            "filename",
            text="파일명"
        )

        self.file_tree.heading(
            "mcode",
            text="MCode"
        )

        self.file_tree.heading(
            "status",
            text="상태"
        )

        self.file_tree.column(
            "filename",
            width=500,
            anchor="w"
        )

        self.file_tree.column(
            "mcode",
            width=120,
            anchor="center"
        )

        self.file_tree.column(
            "status",
            width=90,
            anchor="center"
        )

        self.file_tree.bind(
            "<Double-1>",
            self.open_review_page
        )

        # ------------------------------------------------------------
        # 3. 엑셀 연동 섹션
        # ------------------------------------------------------------
        self.label_excel = Label(self.root, text="배포 경로 정보 (경로정보.xlsx)", font=("맑은 고딕", 11, "bold"), fg="#1D1D1F", bg="#F5F5F7")
        self.label_excel.pack(anchor="w", padx=24, pady=(5, 6))

        self.excel_frame = Frame(self.root, bg="#F5F5F7")
        self.excel_frame.pack(fill=BOTH, padx=24, pady=(0, 15))

        self.btn_add_excel = Button(self.excel_frame, text="엑셀 선택", command=self.add_excel_file, bg="#8E8E93", fg="white", activebackground="#636366", activeforeground="white", **btn_opts)
        self.btn_add_excel.pack(side=LEFT, padx=(0, 10))

        self.label_excel_path = Label(self.excel_frame, text="선택된 엑셀 파일이 없습니다.", font=("맑은 고딕", 9), fg="#8E8E93", bg="#F5F5F7", anchor="w")
        self.label_excel_path.pack(side=LEFT, fill=BOTH, expand=True)

        # ------------------------------------------------------------
        # 4. 검수 보조 기능 영역
        # ------------------------------------------------------------
        self.review_frame = Frame(self.root, bg="#F5F5F7")
        self.review_frame.pack(fill=BOTH, padx=24, pady=(0, 15))
        
        self.btn_copy_urls = Button(self.review_frame, text="📋 성공 URL 전체 복사", command=self.copy_success_urls_to_clipboard, bg="#E8F2FF", fg="#007AFF", activebackground="#D0E3FF", **btn_opts)
        self.btn_copy_urls.pack(side=LEFT, expand=True, fill=BOTH, padx=(0, 4))
        
        self.btn_open_all_success = Button(self.review_frame, text="🚀 성공 브라우저 일괄 오픈", command=self.open_all_success_pages, bg="#E5F9E7", fg="#34C759", activebackground="#C7F3CC", **btn_opts)
        self.btn_open_all_success.pack(side=LEFT, expand=True, fill=BOTH, padx=(4, 0))

        # ------------------------------------------------------------
        # 5. 실행 버튼 및 하단 실시간 콘솔 로그 창 (가독성 최우선 폰트 적용)
        # ------------------------------------------------------------
        self.btn_start = Button(self.root, text="배포 시작하기", command=self.start_deployment_thread, bg="#34C759", fg="white", font=("맑은 고딕", 11, "bold"), relief="flat", cursor="hand2", pady=8, bd=0)
        self.btn_start.pack(fill=BOTH, padx=24, pady=(0, 14))

        self.label_log = Label(self.root, text="작업 로그", font=("맑은 고딕", 13, "bold"), fg="#111827", bg="#F5F5F7")
        self.label_log.pack(anchor="w", padx=24, pady=(0, 4))

        # 폰트: D2Coding / 맑은 고딕 / monospace 순서로 가독성 및 시인성 우수한 고정폭 폰트 선별 지정
        log_font = ("D2Coding", 10)
        self.log_text = Text(
            self.root,
            height=11,
            font=("Consolas", 10),
            bg="#111827",
            fg="#E5E7EB",
            insertbackground="#FFFFFF",
            relief="flat",
            bd=0,
            padx=16,
            pady=14,
            spacing1=2,
            spacing2=1,
            spacing3=2
        )
        self.log_text.pack(
            fill=BOTH,
            expand=True,
            padx=24,
            pady=(6,24)
        )
        self.log_text.config(state=DISABLED)

    def initialize_config_and_logs(self):
        if not os.path.exists("logs"):
            os.makedirs("logs")

        timestamp = datetime.now().strftime("%Y-%m-%d")
        self.log_file_path = f"logs/{timestamp}.log"

        if not os.path.exists(CONFIG_FILE):
            default_config = {
                "excel_path": "",
                "ftp_user": "chunjae\\khr",
                "ftp_pass": "rlagPfla1!"
            }
            try:
                with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                    json.dump(default_config, f, ensure_ascii=False, indent=4)
                self.write_log("INFO", "config.json 파일이 존재하지 않아 기본 파일로 자동 생성했습니다.")
            except Exception as e:
                print(f"초기 설정 파일 생성 오류: {e}")
                
        self.load_config()

    def write_log(self, level, message):
        full_msg = f"[{level}] {message}"
        self.root.after(0, lambda: self._safe_write_text_ui(full_msg))

        if self.log_file_path:
            try:
                with open(self.log_file_path, "a", encoding="utf-8") as f:
                    f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {full_msg}\n")
            except Exception:
                pass

    def _safe_write_text_ui(self, text):
        self.log_text.config(state=NORMAL)
        self.log_text.insert(END, text + "\n")
        self.log_text.see(END)
        self.log_text.config(state=DISABLED)

    # ------------------------------------------------------------
    # FTP 비즈니스 네트워크 통신 코어 (재접속 및 딥 메이크 폴더 탑재)
    # ------------------------------------------------------------

    def ftp_connect(self):
        if not self.ftp_user or not self.ftp_pass:
            self.write_log("ERROR", "config.json 파일에 FTP 계정 정보가 누락되었습니다.")
            return None
        try:
            ftp = FTP()
            # 타임아웃 120초 유지
            ftp.connect(FTP_HOST, FTP_PORT, timeout=120)
            ftp.login(self.ftp_user, self.ftp_pass)

            # [복구] 내 PC 방화벽 블록을 피하기 위해 Passive 모드로 다시 변경
            ftp.set_pasv(True)
            
            ftp.encoding = "utf-8"

            self.remote_root = ""

            try:
                ftp.cwd("/flash")
            except Exception:
                try:
                    ftp.cwd("/Cache/flash")
                    self.remote_root = "/Cache"
                except Exception:
                    raise Exception("FTP 시작 경로를 찾을 수 없습니다. (/flash 또는 /Cache/flash)")

            ftp.cwd("/")
            self.write_log("INFO", f"자동 감지된 시작 경로 : {self.remote_root or '/'}")

            return ftp
        except Exception as e:
            self.write_log("ERROR", f"FTP 최초 연결 에러 ({FTP_HOST}:{FTP_PORT}): {e}")
            return None

    def execute_ftp_with_retry(self, ftp_session, action_func, *args, **kwargs):
        try:
            return action_func(ftp_session, *args, **kwargs), ftp_session
        except Exception as first_error:
            self.write_log("WARNING", f"원격 명령 수행 에러 발생 ({first_error}). 세션을 복구하여 1회 재시도합니다.")
            self.write_log("ERROR", traceback.format_exc())
            
            try:
                ftp_session.close()
            except Exception:
                pass
                
            new_session = self.ftp_connect()
            if not new_session:
                raise Exception(f"네트워크 단절로 인한 FTP 재접속 최종 실패 (상세사유: {first_error})")
                
            try:
                result = action_func(new_session, *args, **kwargs)
                return result, new_session
            except Exception as second_error:
                raise Exception(f"재접속 후 최종 가동 실패 (최종사유: {second_error})")

    def _make_directory_p(self, ftp, remote_dir_path):
        parts = [p for p in remote_dir_path.replace("\\", "/").split("/") if p]
        current_dir = ""
        if remote_dir_path.startswith("/"):
            current_dir = "/"

        for part in parts:
            current_dir = os.path.join(current_dir, part).replace("\\", "/")
            try:
                ftp.cwd(current_dir)
            except error_perm:
                try:
                    ftp.mkd(current_dir)
                    ftp.cwd(current_dir)
                except Exception as e:
                    raise Exception(f"원격 mkdir -p 폴더 생성 오류 ({current_dir}): {e}")

    # ------------------------------------------------------------
    # FTP 원격 입출력 서브 태스크 모듈
    # ------------------------------------------------------------

    def extract_zip(self, zip_path, extract_to):
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(extract_to)
        return True

    def _core_upload_zip(self, ftp, local_zip_path, remote_target_dir, filename):
        self._make_directory_p(ftp, remote_target_dir)
        ftp.cwd(remote_target_dir)
        file_handle = open(local_zip_path, 'rb')
        try:
            ftp.storbinary(f'STOR {filename}', file_handle)
        finally:
            file_handle.close()

    def _core_upload_directory_recursive(self, ftp, local_dir_path, remote_dir_path):
        import time

        for root, dirs, files in os.walk(local_dir_path):

            relative = os.path.relpath(root, local_dir_path)

            if relative == ".":
                current_remote = remote_dir_path
            else:
                current_remote = remote_dir_path.rstrip("/") + "/" + relative.replace("\\", "/")

            self._make_directory_p(ftp, current_remote)
            ftp.cwd(current_remote)

            for filename in files:
                local_file = os.path.join(root, filename)

                # [핵심] 연결 생존 확인 및 끊어졌을 때 자동 재연결
                try:
                    ftp.voidcmd("NOOP")
                except Exception:
                    self.write_log("WARNING", "FTP 연결 끊김 감지, 세션을 재연결합니다...")
                    new_ftp = self.ftp_connect()
                    if new_ftp:
                        ftp = new_ftp
                        ftp.cwd(current_remote)

                self.write_log("INFO", f"업로드 중 : {current_remote}/{filename}")

                # 파일 전송 및 실패 시 1회 즉시 재시도 로직
                try:
                    with open(local_file, "rb") as fp:
                        ftp.storbinary(
                            f"STOR {filename}",
                            fp,
                            blocksize=1024 * 1024,
                        )
                except Exception as e:
                    self.write_log("WARNING", f"업로드 순간 오류 발생({e}), 재연결 후 2차 시도합니다.")
                    ftp = self.ftp_connect()
                    if ftp:
                        ftp.cwd(current_remote)
                        with open(local_file, "rb") as fp:
                            ftp.storbinary(
                                f"STOR {filename}",
                                fp,
                                blocksize=1024 * 1024,
                            )

                self.write_log("INFO", f"업로드 완료 : {current_remote}/{filename}")
                time.sleep(0.05)

    # ------------------------------------------------------------
    # 배포 백그라운드 스레드 제어부
    # ------------------------------------------------------------

    def start_deployment_thread(self):
        if not self.selected_files:
            messagebox.showwarning("경고", "배포할 ZIP 파일을 먼저 추가해 주세요.")
            return
        if not self.excel_path:
            messagebox.showwarning("경고", "경로정보 엑셀 파일을 먼저 선택해 주세요.")
            return

        self.btn_start.config(state=DISABLED)
        
        deploy_thread = threading.Thread(target=self.start_deployment, daemon=True)
        deploy_thread.start()

    def start_deployment(self):
        self.write_log("INFO", "▶ [정밀 webflash 치환 양방향 배포] 프로세스를 개시합니다.\n" + "═"*40)
        
        if not self.find_mcode_from_excel():
            self.write_log("ERROR", "❌ 엑셀 매칭 단계에 치명적 오류가 확인되어 작업을 중단합니다.")
            self.root.after(0, lambda: self.btn_start.config(state=NORMAL))
            return

        ftp = self.ftp_connect()
        if not ftp:
            self.write_log("ERROR", "❌ 초기 FTP 서버 접속에 실패하여 원격 전송 파이프라인을 취소합니다.")
            self.root.after(0, lambda: self.btn_start.config(state=NORMAL))
            return

        success_list = []
        failed_list = []
        
        review_urls_data = []
        csv_report_data = []

        for idx, file_obj in enumerate(self.selected_files):
            filename = file_obj["filename"]
            local_path = file_obj["path"]
            
            # 엑셀 기반 파싱 경로 (기본 구조는 flash 계열)
            parsed_sub_path = file_obj.get("parsed_sub_path", "") # 예: "flash/Multiplayer/A11/N/2025/COE/32"
            
            temp_extract_dir = local_path + "_temp_dir"
            current_time_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            if file_obj["status"] != "매칭완료":
                self.write_log("WARNING", f"⏩ [스킵] {filename} - 매칭 결과가 누락되어 배포 라인업에서 제외됩니다.")
                continue

            progress_text = f"[{idx + 1} / {len(self.selected_files)}] {filename} 정밀 변환 배포 중..."
            self.root.after(0, lambda p=progress_text: self.label_progress.config(text=p, fg="#0066cc"))

            try:
                # ------------------------------------------------------------
                # [트랙 1] 원본 ZIP 파일 배포 (기본 flash 경로 그대로 사용)
                # ------------------------------------------------------------
                target_zip_remote_dir = self.remote_root + "/" + parsed_sub_path.strip("/")
                
                self.update_status(filename, "ZIP업로드")
                self.write_log("INFO", f"📦 [1/2] {filename} ➔ ZIP 전송지: {target_zip_remote_dir}")
                _, ftp = self.execute_ftp_with_retry(ftp, self._core_upload_zip, local_path, target_zip_remote_dir, filename)

                # ------------------------------------------------------------
                # [트랙 2] WEB용 압축 해제 폴더 배포 (flash ➔ webflash 로 중간 단어 치환)
                # ------------------------------------------------------------
                self.update_status(filename, "폴더업로드")
                
                # 대소문자 무관하게 가장 앞단에 출현하는 flash 경로를 webflash로 치환 처리
                web_sub_path = parsed_sub_path.replace("flash", "webflash").replace("Flash", "webflash")
                target_web_base_dir = self.remote_root + "/" + web_sub_path.strip("/")
                
                target_folder_name, _ = os.path.splitext(filename)
                target_folder_remote_path = target_web_base_dir.rstrip('/') + '/' + target_folder_name

                self.write_log("INFO", f"📂 [2/2] {filename} ➔ WEB 폴더 전송지: {target_folder_remote_path}")
                self.extract_zip(local_path, temp_extract_dir)
                
                # 원격지 자동 생성 및 웹 소스 주입
                _, ftp = self.execute_ftp_with_retry(ftp, self._core_upload_directory_recursive, temp_extract_dir, target_folder_remote_path)

                # 전체 양방향 프로세스 성공 완료 처리
                self.update_status(filename, "완료")
                success_list.append(filename)
                
                generated_url = f"https://test-web.milkt.co.kr/HtmlPlayer?id={file_obj['mcode']}"
                review_urls_data.append(generated_url)
                
                csv_report_data.append([filename, file_obj['mcode'], "성공", "", current_time_str])
                self.write_log("INFO", f"✔ [성공] {filename} ➔ (양방향 크로스 매칭 주입 완료)")

            except Exception as file_error:
                self.update_status(filename, "실패")
                error_message = str(file_error)
                failed_list.append((filename, error_message))
                
                csv_report_data.append([filename, file_obj['mcode'], "실패", error_message, current_time_str])
                self.write_log("ERROR", f"❌ [실패] {filename} - 배포 오류 원인: {error_message}")

            finally:
                if os.path.exists(temp_extract_dir):
                    try:
                        shutil.rmtree(temp_extract_dir)
                    except Exception as e:
                        print(f"임시 폴더 삭제 에러: {e}")

            current_percentage = int(((idx + 1) / len(self.selected_files)) * 100)
            self.root.after(0, lambda v=current_percentage: self.progress_bar.config(value=v))

        if ftp:
            try:
                ftp.quit()
            except Exception:
                try:
                    ftp.close()
                except Exception:
                    pass

        # 종합 리포트 출력 마감
        self.root.after(0, lambda: self.label_progress.config(text="모든 양방향 크로스 배포 작업이 마무리되었습니다.", fg="#27ae60"))
        
        if review_urls_data:
            try:
                with open(REVIEW_FILE, "w", encoding="utf-8") as rf:
                    rf.write("\n".join(review_urls_data))
                self.write_log("INFO", f"💾 시스템: 성공 콘텐츠 검수 URL 목록이 '{REVIEW_FILE}'로 자동 생성되었습니다.")
                
                self.root.clipboard_clear()
                self.root.clipboard_append("\n".join(review_urls_data))
                self.root.update()
                self.write_log("INFO", "📋 시스템: 성공 건에 대한 모든 검수 URL이 클립보드에 자동 복사되었습니다.")
            except Exception as fe:
                self.write_log("ERROR", f"⚠️ 검수 URL 텍스트 파일 저장 중 예외 발생: {fe}")

        try:
            with open(RESULT_CSV_FILE, "w", encoding="utf-8-sig", newline="") as cf:
                writer = csv.writer(cf)
                writer.writerow(["파일명", "mcode", "배포결과", "실패사유", "배포시간"])
                writer.writerows(csv_report_data)
            self.write_log("INFO", f"💾 시스템: 종합 배포 결과 명세서가 '{RESULT_CSV_FILE}'로 내보내졌습니다.")
        except Exception as ce:
            self.write_log("ERROR", f"⚠️ 종합 결과 분석 CSV 파일 생성 실패: {ce}")
        
        summary_report = [
            "\n================================",
            "정밀 양방향 크로스 배포 완료",
            f"\n성공 개수 : {len(success_list)}건",
            f"실패 개수 : {len(failed_list)}건",
            "\n실패 사유"
        ]
        if failed_list:
            for item, reason in failed_list:
                summary_report.append(f"- {item} ➔ {reason}")
        else:
            summary_report.append("- 없음")
        summary_report.append("================================")
        
        self.write_log("INFO", "\n".join(summary_report))
        self.root.after(0, lambda: self.btn_start.config(state=NORMAL))

    # ------------------------------------------------------------
    # 사후 검수 부가 기능 이벤트 핸들러
    # ------------------------------------------------------------

    def copy_success_urls_to_clipboard(self):
        success_urls = []
        for file_obj in self.selected_files:
            if file_obj["status"] == "완료" and file_obj["mcode"]:
                url = f"https://test-web.milkt.co.kr/HtmlPlayer?id={file_obj['mcode']}"
                success_urls.append(url)

        if not success_urls:
            messagebox.showwarning("안내", "배포 완료(성공) 상태의 유효한 mcode 콘텐츠가 없습니다.")
            return

        self.root.clipboard_clear()
        self.root.clipboard_append("\n".join(success_urls))
        self.root.update()
        messagebox.showinfo("복사 완료", "성공한 콘텐츠의 검수 URL이 클립보드에 일괄 복사되었습니다.")

    def open_all_success_pages(self):
        success_count = 0
        for file_obj in self.selected_files:
            if file_obj["status"] == "완료" and file_obj["mcode"]:
                url = f"https://test-web.milkt.co.kr/HtmlPlayer?id={file_obj['mcode']}"
                try:
                    webbrowser.open_new_tab(url)
                    success_count += 1
                except Exception as e:
                    self.write_log("WARNING", f"브라우저 호출 중 예외 발생: {e}")

        if success_count == 0:
            messagebox.showwarning("안내", "일괄 오픈할 배포 성공 완료 항목이 목록에 존재하지 않습니다.")

    # ------------------------------------------------------------
    # UI 이벤트 및 유틸리티 함수
    # ------------------------------------------------------------

    def open_review_page(self, event=None):
        selected_items = self.file_tree.selection()
        if not selected_items:
            messagebox.showinfo("안내", "검수할 콘텐츠를 목록에서 선택해주세요.")
            return

        for item in selected_items:
            values = self.file_tree.item(item, "values")
            filename, mcode = values[0], values[1]

            if not mcode or mcode.strip() == "":
                self.write_log("WARNING", f"❌ [오류] 먼저 mcode를 조회해주세요. (파일명: {filename})")
                continue

            try:
                review_url = f"https://test-web.milkt.co.kr/HtmlPlayer?id={mcode}"
                webbrowser.open_new_tab(review_url)
            except Exception as e:
                self.write_log("ERROR", f"브라우저 연동 실패: {e}")

    def add_zip_files(self):
        try:
            files = filedialog.askopenfilenames(title="배포할 ZIP 파일들을 선택하세요", filetypes=[("ZIP 파일", "*.zip"), ("모든 파일", "*.*")])
            if files:
                for file_path in files:
                    existing_paths = [file_obj["path"] for file_obj in self.selected_files]
                    if file_path not in existing_paths:
                        file_name = os.path.basename(file_path)
                        file_object = {"path": file_path, "filename": file_name, "mcode": "", "status": "대기", "parsed_sub_path": ""}
                        self.selected_files.append(file_object)
                        self.file_tree.insert("", END, values=(file_name, "", "대기"))
                self.write_log("INFO", f"ZIP 파일 {len(files)}개가 목록에 추가되었습니다.")
        except Exception as e:
            messagebox.showerror("오류", f"파일 시스템 스캔 실패:\n{e}")

    def update_status(self, filename, new_status):
        for file_obj in self.selected_files:
            if file_obj["filename"] == filename:
                file_obj["status"] = new_status
                break
        self.root.after(0, lambda: self._safe_update_tree_status(filename, new_status))

    def _safe_update_tree_status(self, filename, new_status):
        for item in self.file_tree.get_children():
            values = self.file_tree.item(item, "values")
            if values[0] == filename:
                self.file_tree.item(item, values=(values[0], values[1], new_status))
                break

    def update_mcode_and_path(self, filename, mcode, parsed_sub_path):
        for file_obj in self.selected_files:
            if file_obj["filename"] == filename:
                file_obj["mcode"] = mcode
                file_obj["parsed_sub_path"] = parsed_sub_path
                break
        self.root.after(0, lambda: self._safe_update_tree_mcode(filename, mcode))

    def _safe_update_tree_mcode(self, filename, mcode):
        for item in self.file_tree.get_children():
            values = self.file_tree.item(item, "values")
            if values[0] == filename:
                self.file_tree.item(item, values=(values[0], mcode, values[2]))
                break

    def get_filename_from_path(self, path_string):
        if not path_string: return ""
        return path_string.replace("\\", "/").split("/")[-1]

    def parse_excel_route_to_sub_dir(self, full_route_str):
        clean_path = full_route_str.replace("\\", "/").strip()
        if not clean_path:
            return ""
        
        parts = [p for p in clean_path.split("/") if p]
        if not parts:
            return ""
            
        if "." in parts[-1]:
            parts = parts[:-1]
            
        return "/".join(parts)

    def find_mcode_from_excel(self):
        wb = None
        try:
            wb = openpyxl.load_workbook(self.excel_path, data_only=True)
            sheet = wb.active 
            mcode_col_idx, path_col_idx = None, None
            
            for col_idx in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col_idx).value
                if cell_value == "mcode": mcode_col_idx = col_idx
                elif cell_value == "경로": path_col_idx = col_idx

            if mcode_col_idx is None or path_col_idx is None:
                self.write_log("WARNING", "⚠️ [형식오류] 엑셀에 'mcode' 또는 '경로' 컬럼명이 존재하지 않습니다.")
                return False

            for file_obj in self.selected_files:
                target_zip_name = file_obj["filename"] 
                is_matched = False 
                
                for row_idx in range(2, sheet.max_row + 1):
                    excel_path_value = sheet.cell(row=row_idx, column=path_col_idx).value
                    excel_mcode_value = sheet.cell(row=row_idx, column=mcode_col_idx).value
                    
                    if excel_path_value:
                        excel_filename = self.get_filename_from_path(str(excel_path_value))
                        if excel_filename.lower() == target_zip_name.lower():
                            mcode_result = str(excel_mcode_value) if excel_mcode_value else ""
                            
                            extracted_sub_dir = self.parse_excel_route_to_sub_dir(str(excel_path_value))
                            
                            self.update_mcode_and_path(target_zip_name, mcode_result, extracted_sub_dir)
                            self.update_status(target_zip_name, "매칭완료")
                            is_matched = True
                            break 
                
                if not is_matched:
                    self.update_status(target_zip_name, "경로없음")
                    self.write_log("WARNING", f"[경로없음] 엑셀 매칭 정보 없음 - ZIP: {target_zip_name}")
            return True
        except Exception as e:
            self.write_log("ERROR", f"⚠️ [엑셀오류] 파일 로드 실패: {e}")
            return False
        finally:
            if wb:
                try:
                    wb.close()
                except Exception:
                    pass

    def delete_selected_files(self):
        selected_items = self.file_tree.selection()
        if not selected_items: return
        for item in selected_items:
            values = self.file_tree.item(item, "values")
            file_name = values[0] 
            for file_obj in self.selected_files:
                if file_obj["filename"] == file_name:
                    self.selected_files.remove(file_obj)
                    break
            self.file_tree.delete(item)

    def delete_all_files(self):
        if not self.selected_files: return
        if messagebox.askyesno("확인", "목록에 있는 모든 파일을 삭제하시겠습니까?"):
            self.selected_files.clear() 
            for item in self.file_tree.get_children(): self.file_tree.delete(item)
            self.write_log("INFO", "모든 파일 목록이 초기화되었습니다.")

    def add_excel_file(self):
        try:
            file = filedialog.askopenfilename(title="경로정보 엑셀 파일을 선택하세요", filetypes=[("엑셀 파일", "*.xlsx *.xls"), ("모든 파일", "*.*")])
            if file:
                self.excel_path = file
                self.label_excel_path.config(text=file, fg="#0066cc")
                self.write_log("INFO", f"연동용 경로 엑셀 지정 완료: {os.path.basename(file)}")
                self.save_config()
        except Exception as e:
            messagebox.showerror("오류", f"엑셀 선택 차단 예외: {e}")

    def save_config(self):
        try:
            config_data = {"excel_path": self.excel_path, "ftp_user": self.ftp_user, "ftp_pass": self.ftp_pass}
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(config_data, f, ensure_ascii=False, indent=4)
        except Exception as e:
            print(f"설정 입출력 에러: {e}")

    def load_config(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    config_data = json.load(f)
                    self.excel_path = config_data.get("excel_path", "")
                    self.ftp_user = config_data.get("ftp_user", "")
                    self.ftp_pass = config_data.get("ftp_pass", "")
                    if self.excel_path and os.path.exists(self.excel_path):
                        self.label_excel_path.config(text=self.excel_path, fg="#0066cc")
            except Exception as e:
                print(f"설정 복원 에러: {e}")

if __name__ == "__main__":
    root = Tk()
    app = DeploymentToolApp(root)
    root.mainloop()